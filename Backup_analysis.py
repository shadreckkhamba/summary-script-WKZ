import os
import calendar
import gzip
import shutil
import mysql.connector
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Configuration
DB_HOST = "localhost"
DB_USER = "root"
DB_PASSWORD = "password"
TEMP_DB = "billing_analysis"
BACKUP_DIR = "/home/shadreck/Documents/backup"

current_date = datetime.now()
start_date = current_date - timedelta(days=30)
start_date_str = start_date.strftime('%Y-%m-%d')
end_date_str = current_date.strftime('%Y-%m-%d')

backup_files = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".sql.gz")], reverse=True)
if not backup_files:
    print("No backup files found.")
    exit(1)
backup_file = os.path.join(BACKUP_DIR, backup_files[0])
print(f"Using backup file: {backup_file}")

# Extract the SQL file
temp_sql_file = backup_file.replace(".gz", "")
with gzip.open(backup_file, 'rb') as f_in:
    with open(temp_sql_file, 'wb') as f_out:
        shutil.copyfileobj(f_in, f_out)
print(f"Extracted SQL file: {temp_sql_file}")

conn = mysql.connector.connect(host=DB_HOST, user=DB_USER, password=DB_PASSWORD)
cursor = conn.cursor()
cursor.execute(f"DROP DATABASE IF EXISTS {TEMP_DB}")
cursor.execute(f"CREATE DATABASE {TEMP_DB}")
cursor.close()
conn.close()
os.system(f"mysql -u {DB_USER} -p{DB_PASSWORD} {TEMP_DB} < {temp_sql_file}")
print("Database restored successfully.")

conn = mysql.connector.connect(host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=TEMP_DB)
cursor = conn.cursor()


# 1. Registered Patients Summary
registered_patients_query = """
SELECT 
    (SELECT COUNT(*) 
     FROM patient 
     WHERE voided = 0 
     AND date_created >= MAKEDATE(YEAR(CURDATE()), 1)) AS this_year,
     
    (SELECT COUNT(*) 
     FROM patient 
     WHERE voided = 0 
     AND date_created >= DATE_FORMAT(CURDATE(), '%Y-%m-01')) AS this_month,
     
    (SELECT COUNT(*) 
     FROM patient 
     WHERE voided = 0 
     AND date_created >= DATE_SUB(CURDATE(), INTERVAL WEEKDAY(CURDATE()) DAY)) AS this_week,
     
    (SELECT COUNT(*) 
     FROM patient 
     WHERE voided = 0 
     AND DATE(date_created) = CURDATE()) AS today;
"""
cursor.execute(registered_patients_query)
this_year, this_month, this_week, today = cursor.fetchone()

# Paying vs Non-Paying Query
paying_query = """
SELECT 
    COUNT(*) AS total_patients,
    SUM(CASE WHEN paying_orders > 0 AND non_paying_orders = 0 THEN 1 ELSE 0 END) AS exclusively_paying,
    SUM(CASE WHEN non_paying_orders > 0 AND paying_orders = 0 THEN 1 ELSE 0 END) AS exclusively_non_paying,
    SUM(CASE WHEN paying_orders > 0 AND non_paying_orders > 0 THEN 1 ELSE 0 END) AS both_categories
FROM (
    SELECT 
        patient_id,
        SUM(CASE WHEN full_price >= 1000 THEN 1 ELSE 0 END) AS paying_orders,
        SUM(CASE WHEN full_price = 0 THEN 1 ELSE 0 END) AS non_paying_orders
    FROM order_entries
    WHERE order_date BETWEEN %s AND %s
    GROUP BY patient_id
) AS patient_summary;
"""
cursor.execute(paying_query, (f"{start_date_str} 00:00:00", f"{end_date_str} 23:59:59"))
total, paying, non_paying, both = cursor.fetchone()

# Combine the first part only in DataFrame
registered_patients_df = pd.DataFrame({
    "Metric": ["Registered This Year", "Registered This Month", "Registered This Week", "Registered Today"],
    "Count": [this_year, this_month, this_week, today]
})

# 2. Order Entries Analysis
order_entries_query = """
    SELECT s.service_id, 
           s.name AS service_name, 
           SUM(oe.quantity) AS total_quantity, 
           SUM(oe.amount_paid) AS total_amount_paid,
           SUM(oe.quantity * sp.price) AS expected_total_amount_paid,
           COUNT(DISTINCT CASE WHEN oe.amount_paid < oe.full_price THEN oe.patient_id END) AS patients_with_balance
    FROM order_entries oe
    JOIN services s ON oe.service_id = s.service_id
    JOIN service_prices sp ON oe.service_id = sp.service_id AND sp.voided = 0
    WHERE oe.voided = 0
    GROUP BY s.service_id, s.name;
"""
cursor.execute(order_entries_query)
order_entries_results = cursor.fetchall()

order_entries_df = pd.DataFrame(order_entries_results, columns=[
    "  Service ID   ", "Service Name", "   Total Quantity   ", "Total Amount Paid", 
    "Expected Total Amount Paid", "Patients With Outstanding Balance"
])

order_entries_df["Total Amount Overdue"] = order_entries_df["Expected Total Amount Paid"] - order_entries_df["Total Amount Paid"]

total_quantity = order_entries_df["   Total Quantity   "].sum()
total_amount_collected = order_entries_df["Total Amount Paid"].sum()
total_expected = order_entries_df["Expected Total Amount Paid"].sum()
total_overdue = order_entries_df["Total Amount Overdue"].sum()
total_patients_balance = order_entries_df["Patients With Outstanding Balance"].sum()

totals_df = pd.DataFrame([{
    "Total Amount Paid": total_amount_collected,
    "Expected Total Amount Paid": total_expected,
    "Patients With Outstanding Balance": total_patients_balance,
    "Total Amount Overdue": total_overdue
}])

order_entries_df = pd.concat([order_entries_df, totals_df], ignore_index=True)

for col in ["Total Amount Paid", "Expected Total Amount Paid", "Total Amount Overdue"]:
    order_entries_df[col] = order_entries_df[col].apply(lambda x: f"MWK {x:,.2f}" if isinstance(x, (int, float)) else x)


# 3. Patient Age Groups
age_group_query = """
SELECT age_group, 
       gender,
       COUNT(*) AS total_patients
FROM (
    SELECT 
        CASE 
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, NOW()) < 5 THEN 'Under 5'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, NOW()) BETWEEN 5 AND 9 THEN '5-9'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, NOW()) BETWEEN 10 AND 14 THEN '10-14'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, NOW()) BETWEEN 15 AND 19 THEN '15-19'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, NOW()) BETWEEN 20 AND 24 THEN '20-24'
            ELSE 'Other'
        END AS age_group, 
        p.gender
    FROM patient pat
    JOIN person p ON pat.patient_id = p.person_id
    WHERE pat.date_created BETWEEN DATE_SUB(NOW(), INTERVAL 30 DAY) AND NOW()
) AS subquery
GROUP BY age_group, gender
ORDER BY age_group, gender;
"""
cursor.execute(age_group_query)
age_group_results = cursor.fetchall()

age_group_df = pd.DataFrame(age_group_results, columns=["Age Group", "Gender", "Total Patients"])
age_group_df["Age Group"] = pd.Categorical(age_group_df["Age Group"], categories=['Under 5', '5-9', '10-14', '15-19', '20-24', 'Other'], ordered=True)
age_group_df = age_group_df.sort_values(["Age Group", "Gender"])


# 4. Services Used Per Age Group
most_profitable_services_query = """
    SELECT 
        CASE 
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, CURDATE()) < 5 THEN 'Under 5'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, CURDATE()) BETWEEN 6 AND 17 THEN '6-17'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, CURDATE()) BETWEEN 18 AND 35 THEN '18-35'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, CURDATE()) BETWEEN 36 AND 50 THEN '36-50'
            WHEN TIMESTAMPDIFF(YEAR, p.birthdate, CURDATE()) > 50 THEN 'Above 50'
            ELSE 'Unknown'
        END AS age_group,
        s.name AS service_name, 
        SUM(oe.amount_paid) AS total_amount_paid
    FROM order_entries oe
    JOIN services s ON oe.service_id = s.service_id
    JOIN patient pt ON oe.patient_id = pt.patient_id
    JOIN person p ON pt.patient_id = p.person_id
    WHERE oe.voided = 0
    GROUP BY age_group, service_name
    ORDER BY total_amount_paid DESC;
"""
cursor.execute(most_profitable_services_query)
most_profitable_services_results = cursor.fetchall()
most_profitable_services_df = pd.DataFrame(most_profitable_services_results, columns=["Age Group", "Service Name", "Total Amount Paid"])
most_profitable_services_df["Total Amount Paid"] = most_profitable_services_df["Total Amount Paid"].apply(lambda x: f"MWK {x:,.2f}")

# 5. Most Popular Services Overall
most_popular_services_query = """
    SELECT 
        s.name AS service_name, 
        SUM(oe.quantity) AS total_quantity,
        SUM(oe.amount_paid) AS total_amount_paid,
        sp.price AS service_price,
        sp.price_type AS price_type
    FROM order_entries oe
    JOIN services s ON oe.service_id = s.service_id
    JOIN service_prices sp ON s.service_id = sp.service_id AND sp.voided = 0
    WHERE oe.voided = 0
    GROUP BY service_name, sp.price, sp.price_type
    ORDER BY total_quantity DESC;
"""
cursor.execute(most_popular_services_query)
most_popular_services_results = cursor.fetchall()

most_popular_services_df = pd.DataFrame(most_popular_services_results, 
    columns=["Service Name", "Total Quantity", "Total Amount Paid", "Service Price", "Price Type"])

# 6. Services used per month
services_used_per_month_query = """
    SELECT 
        s.name AS service_name,
        YEAR(oe.order_date) AS year,
        MONTH(oe.order_date) AS month,
        COUNT(*) AS services_used_per_month
    FROM order_entries oe
    JOIN services s ON oe.service_id = s.service_id
    WHERE oe.voided = 0
    GROUP BY service_name, year, month
    ORDER BY year DESC, month DESC;
"""
cursor.execute(services_used_per_month_query)
services_used_results = cursor.fetchall()

services_used_results = [
    (service_name, year, calendar.month_name[month], services_used) 
    for service_name, year, month, services_used in services_used_results
    ]
services_used_df = pd.DataFrame(services_used_results, columns=["Service Name", "Year", "Month", "Services Used Per Month"])
services_used_df["Month_num"] = services_used_df["Month"].map({month: idx for idx, month in enumerate(calendar.month_name) if month})
services_used_df.sort_values(by=["Year", "Month_num"], ascending=[False, False], inplace=True)
services_used_df.drop("Month_num", axis=1, inplace=True)


#7. Get distribution of returning patients
returning_patients_query = f"""
SELECT COUNT(*) 
FROM (
    SELECT patient_id 
    FROM receipts 
    WHERE payment_stamp BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
    GROUP BY patient_id 
    HAVING COUNT(receipt_number) > 1
) AS subquery;
"""
cursor.execute(returning_patients_query)
returning_patients_count = cursor.fetchone()[0]

#8 Patient distribution based on age and gender
returning_patients_distribution_query = f"""
SELECT 
    CASE 
        WHEN TIMESTAMPDIFF(YEAR, per.birthdate, CURDATE()) < 5 THEN 'under_five'
        WHEN TIMESTAMPDIFF(YEAR, per.birthdate, CURDATE()) BETWEEN 5 AND 12 THEN 'under_thirteen'
        ELSE 'adult'
    END AS age_category,
    per.gender AS gender,
    COUNT(DISTINCT r.patient_id) AS returning_patient_count
FROM receipts r
JOIN patient p ON r.patient_id = p.patient_id
JOIN person per ON p.patient_id = per.person_id
WHERE r.payment_stamp BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
  AND r.patient_id IN (
    SELECT patient_id 
    FROM receipts
    WHERE payment_stamp BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
    GROUP BY patient_id
    HAVING COUNT(*) > 1
)
GROUP BY age_category, per.gender
ORDER BY age_category, per.gender;
"""
cursor.execute(returning_patients_distribution_query)
returning_patients_results = cursor.fetchall()

empty_row = pd.DataFrame([[""] * len(order_entries_df.columns)], columns=order_entries_df.columns)
for _ in range(1):
    order_entries_df = pd.concat([order_entries_df, empty_row], ignore_index=True)

order_entries_df = pd.concat([order_entries_df, pd.DataFrame([{
    "Service Name": "",
    "Total Amount Paid": "",
    "Expected Total Amount Paid": "",
    "Patients With Outstanding Balance": ""
}])], ignore_index=True)

distribution_data = []
for age_category, gender, count in returning_patients_results:
    distribution_data.append({
        "Service Name": "",
        "Total Amount Paid": "",
        "Expected Total Amount Paid": "",
        "Patients With Outstanding Balance": ""
    })

distribution_df = pd.DataFrame(distribution_data)
order_entries_df = pd.concat([order_entries_df, distribution_df], ignore_index=True)

# Returning patients frequency
returning_patients_frequency_query = f"""
SELECT
    visit_count AS 'Number of Visits',
    COUNT(patient_id) AS 'Number of Patients'
FROM (
    SELECT
        patient_id,
        COUNT(*) AS visit_count
    FROM receipts
    WHERE payment_stamp BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
    GROUP BY patient_id
    HAVING COUNT(*) > 1
) AS returning_patient_visits
GROUP BY visit_count
ORDER BY visit_count;
"""
cursor.execute(returning_patients_frequency_query)
returning_patients_freq_results = cursor.fetchall()

#9. SQL query for the trend of money made per day
trend_query = f"""
SELECT 
    DATE(created_at) AS transaction_date,
    SUM(full_price) AS total_collected
FROM order_entries
WHERE cashier IN ('1','8', '9')
AND created_at BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
GROUP BY transaction_date
ORDER BY transaction_date;
"""
cursor.execute(trend_query)
result = cursor.fetchall()
if result:
    trend_df = pd.DataFrame(result)
    trend_df.columns = ["Transaction Date", "Total Collected"]
    trend_df["Transaction Date"] = pd.to_datetime(trend_df["Transaction Date"])
    trend_df["Total Collected"] = trend_df["Total Collected"].apply(
        lambda x: f"MWK {x:,.2f}" if isinstance(x, (int, float)) else x
    )
    trend_df["Transaction Date"] = trend_df["Transaction Date"].dt.strftime("%Y-%m-%d")
else:
    trend_df = pd.DataFrame(columns=["Transaction Date", "Total Collected"])


#10. SQL query for Daily Hospital Patient Visits
hospital_visits_query = f"""
SELECT
    daily_registrations.registration_date,
    daily_registrations.total_registrations,
    COALESCE(daily_returning.total_returning_patients, 0) AS total_returning_patients,
    daily_registrations.total_registrations + COALESCE(daily_returning.total_returning_patients, 0) AS total_visits
FROM (
    SELECT DATE(date_created) AS registration_date, COUNT(*) AS total_registrations
    FROM patient
    WHERE date_created BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
    GROUP BY registration_date
) AS daily_registrations
LEFT JOIN (
    SELECT
        visit_date,
        COUNT(patient_id) AS total_returning_patients
    FROM (
        SELECT
            DATE(payment_stamp) AS visit_date,
            patient_id,
            COUNT(*) AS visit_count
        FROM receipts
        WHERE payment_stamp BETWEEN '{start_date_str} 00:00:00' AND '{end_date_str} 23:59:59'
        GROUP BY visit_date, patient_id
        HAVING COUNT(*) > 1
    ) AS daily_returning
    GROUP BY visit_date
) AS daily_returning ON daily_registrations.registration_date = daily_returning.visit_date
ORDER BY daily_registrations.registration_date;
"""
cursor.execute(hospital_visits_query)
result = cursor.fetchall()
if result:
    hospital_visits_df = pd.DataFrame(result)
    hospital_visits_df.columns = ["Registration Date", "Total Registrations", "Total Returning Patients", "Total Visits"]
    hospital_visits_df["Registration Date"] = pd.to_datetime(hospital_visits_df["Registration Date"])
    hospital_visits_df["Registration Date"] = hospital_visits_df["Registration Date"].dt.strftime("%Y-%m-%d")
else:
    hospital_visits_df = pd.DataFrame(columns=["Registration Date", "Total Registrations", "Total Returning Patients", "Total Visits"])

cursor.close()
conn.close()

consolidated_file = os.path.join(BACKUP_DIR, "Consolidated_Report.xlsx")
with pd.ExcelWriter(consolidated_file, engine='openpyxl') as writer:
    registered_patients_df.to_excel(writer, sheet_name="Registered Patients", index=False)
    order_entries_df.to_excel(writer, sheet_name="Order Entries", index=False)
    age_group_df.to_excel(writer, sheet_name="Registered Patient Age Groups", index=False)
    most_profitable_services_df.to_excel(writer, sheet_name="Service Profits By Age Group", index=False)
    most_popular_services_df.to_excel(writer, sheet_name="Popular Services", index=False)
    services_used_df.to_excel(writer, sheet_name="Services Used Per Month", index=False)
    trend_df.to_excel(writer, sheet_name="Daily Money Trend", index=False)
    hospital_visits_df.to_excel(writer, sheet_name="Daily Hospital Patient Visits", index=False)
pdf_file = os.path.join(BACKUP_DIR, "Consolidated_Report.pdf")

wb = load_workbook(consolidated_file)
ws = wb["Registered Patients"]

right_title = "Paying vs. Non-Paying Patients"
right_data = [
    ("Total Patients", total),
    ("Exclusively Paying Patients", paying),
    ("Exclusively Non-Paying Patients", non_paying),
    ("Patients in Both Categories", both)
]

start_col = 4 
start_row = 1
ws.cell(row=start_row, column=start_col).value = right_title
ws.cell(row=start_row, column=start_col).font = Font(bold=True)

for i, (label, count) in enumerate(right_data, start=1):
    ws.cell(row=start_row + i, column=start_col).value = label
    ws.cell(row=start_row + i, column=start_col + 1).value = count

ws.column_dimensions[get_column_letter(start_col)].width = 35
ws.column_dimensions[get_column_letter(start_col + 1)].width = 15

wb.save(consolidated_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.insert_rows(1)

    ws['A1'] = sheet_name
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    if sheet_name == "Order Entries":
        returning_patients_row = len(order_entries_df)

        ws.merge_cells(start_row=returning_patients_row, start_column=1, end_row=returning_patients_row, end_column=3)
        ws[f"A{returning_patients_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{returning_patients_row}"].font = Font(bold=True)
        ws[f"A{returning_patients_row}"] = f"Returning Patients Distribution · {start_date_str} to {end_date_str}"
        ws.row_dimensions[returning_patients_row].height = 30

        ws[f"A{returning_patients_row + 1}"] = "Distribution"
        ws[f"B{returning_patients_row + 1}"] = "Count"
        ws[f"C{returning_patients_row + 1}"] = "Total Patients"
        ws[f"A{returning_patients_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{returning_patients_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{returning_patients_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{returning_patients_row + 1}"].font = Font(bold=True)
        ws[f"B{returning_patients_row + 1}"].font = Font(bold=True)
        ws[f"C{returning_patients_row + 1}"].font = Font(bold=True)

        total_returning_patients = 0
        for idx, (age_category, gender, count) in enumerate(returning_patients_results, start=returning_patients_row + 2):
            ws[f"A{idx}"] = f"{age_category} ({gender})"
            ws[f"B{idx}"] = count
            ws[f"A{idx}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"B{idx}"].alignment = Alignment(horizontal="center", vertical="center")
            total_returning_patients += count

        total_row = returning_patients_row + 2 + len(returning_patients_results)
        ws[f"B{total_row}"].font = Font(bold=True)
        ws[f"C{total_row}"] = total_returning_patients
        ws[f"A{total_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"C{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{total_row}"].font = Font(bold=True)

        frequency_title_row = total_row + 2
        ws.merge_cells(start_row=frequency_title_row, start_column=1, end_row=frequency_title_row, end_column=3)
        ws[f"A{frequency_title_row}"] = "Frequency of The Returning Patients"
        ws[f"A{frequency_title_row}"].font = Font(bold=True)
        ws[f"A{frequency_title_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[frequency_title_row].height = 30

        freq_header_row = frequency_title_row + 1
        ws[f"A{freq_header_row}"] = "Number of Visits"
        ws[f"B{freq_header_row}"] = "Number of Patients"
        ws[f"C{freq_header_row}"] = "Patients With More Visits"
        ws[f"A{freq_header_row}"].font = Font(bold=False)
        ws[f"B{freq_header_row}"].font = Font(bold=False)
        ws[f"A{freq_header_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{freq_header_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{freq_header_row}"].alignment = Alignment(horizontal="center", vertical="center")

        total_patients_with_more_visits = 0
        for idx_offset, (visits, patient_count) in enumerate(returning_patients_freq_results):
            row = freq_header_row + 1 + idx_offset
            ws[f"A{row}"] = visits
            ws[f"B{row}"] = patient_count
            ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
            total_patients_with_more_visits += patient_count

        final_freq_row = freq_header_row + 1 + len(returning_patients_freq_results)
        ws[f"C{final_freq_row}"] = total_patients_with_more_visits
        ws[f"C{final_freq_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{final_freq_row}"].font = Font(bold=True)

    ws.protection.sheet = True
    ws.protection.password = 'ghii@wkz'

wb.save(consolidated_file)
print(f"Consolidated report saved: {consolidated_file}")

# Send to virtual server
result = os.system(f"scp {consolidated_file} ghii@192.168.10.186:/home/ghii/tests/backup")
if result == 0:
    print("Report sent to virtual server")
else: 
    print("Failed to send the report to virtual server: Lost Connection")